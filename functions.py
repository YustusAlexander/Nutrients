import openpyxl
import re
path_db_file = 'data/usda_full.xlsx'
wb_db = openpyxl.load_workbook(path_db_file)
ws_db = wb_db["db"]

path_calc_file = 'data/list_for_result.xlsx'
wb_calc = openpyxl.load_workbook(path_calc_file)
ws_calc = wb_calc["calc"]

################################## show ############################################

def show_food():
    show_column = [1, 3]
    return "\n".join([" | ".join([str(ws_calc.cell(row=i + 3, column=j).value).lower() for j in show_column])
                      for i in range(ws_calc.max_row - 2)])

def calculate():
    d = {}
    for i in range(1, ws_db.max_column + 1):
        val_sum_mas = 0
        if str(ws_calc.cell(row=1, column=i).value).isdigit():   # проверка наличия нормы
            for j in range(3, ws_db.max_row + 1):  # сумма в столбце
                if str(ws_calc.cell(row=j, column=i).value).isdigit():
                    val_sum_mas += float(ws_calc.cell(row=j, column=i).value)
            val_norm = float(ws_calc.cell(row=1, column=i).value)
            val_percent = (val_sum_mas * 100) / val_norm
            name_percent = ws_calc.cell(row=2, column=i).value
            d[name_percent] = val_percent
    wb_calc.close()
    return d

def show_calculate(dict_calc):
    return "\n".join([str(int(value)) + "%  " + key for key, value in dict_calc.items()])



################################## add ############################################

def find_food(find_text):
    find_row = []
    search_text = find_text.upper()
    column_find = 'B'
    for row in range(2, ws_db.max_row + 1):
        data_from_cell = ws_db[column_find + str(row)].value
        data_from_cell = str(data_from_cell).upper()
        result = re.findall(search_text, data_from_cell)
        if len(result) > 0:
            find_row.append(row)
    return find_row

def show_find_food(find_row):
    if len(find_row) > 0:  # количество найденных продуктов
        return "\n".join([str(i) + ". " + ws_db.cell(row=i, column=2).value.lower() for i in find_row])
    else:
        return False

def save_food(row, weight):
    row = int(row)
    row_write = 3
    text_column = [1, 2]  # №_db, Description
    ws_calc.insert_rows(idx=row_write, amount=1)  # Вставка новой строки в таблицу calc для расчёта значений
    ws_calc["A" + str(row_write)] = weight
    for i in range(1, ws_db.max_row + 1):
        val_db = ws_db.cell(row=row, column=i).value  # Строка из db
        if i not in text_column and str(val_db).isdigit():
            val_calc = float(val_db) * float(weight) / 100  # Вычисляемое значение массы
        else:
            val_calc = val_db  # Текстовое значение
        ws_calc.cell(row=row_write, column=i + 1).value = val_calc
    wb_calc.save(path_calc_file)
    wb_calc.close()




################################## del  ############################################

def show_del_food():
    dispose = 2
    return "\n".join([str(i) + ". " + str(ws_calc.cell(row=i + dispose, column=3).value).lower()
                      for i in range(1, ws_calc.max_row - dispose)])

def del_food(row_del):
    dispose = 2
    ws_calc.delete_rows(idx=row_del + dispose, amount=1)
    wb_calc.close()
